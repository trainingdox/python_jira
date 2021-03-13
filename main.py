import sys

import requests
import urllib
import json
import openpyxl
from pathlib import Path
import win32com.client as win32
from datetime import date

def get_data_of_alt_id(arg_release_name, arg_alt_id):
    release_name=arg_release_name #its a filter for version/release
    alt_id=arg_alt_id   #its the customer field your test management tool uses to filter testcases
    #proxies = {'http':'','https':''}
    #my_headers = {'Authorization': 'Bearer '}

    #get release_id
    release_id=0
    response_1=requests.get('https://**********.com/flex/services/rest/latest/release',headers=my_headers,proxies=proxies)
    if response_1.status_code == 200:
        list_release=response_1.json()
        for release in list_release:
            if release['name'] == release_name:
                release_id=int(release['id'])
        #if release_id != 0:
            #print("Your release_id is - " + str(release_id)+"\n")
        #else:
            #print("This release name does not exist in Zephyr. ")
        if release_id ==0:
            return 1
    #get TC count for each release_id

    url_part_1='https://*********/flex/services/rest/latest/advancesearch?word=altId~"'
    url_part_2='"&entitytype=testexecution&releaseid='
    url_part_3='&zql=true&isascorder=true&firstresult=0&maxresults=2000&isOld=false'
    url = url_part_1 + alt_id + url_part_2 + str(release_id) + url_part_3
    word=altId~"**********"&entitytype=testexecution&releaseid=1167&zql=true&isascorder=true&firstresult=0&maxresults=2000&isOld=false'
    print("Getting details for alt_id - " + alt_id)
    response=requests.get( url, headers=my_headers,proxies=proxies)
    if (response.status_code == 200):
        list1=response.json()
        pass_count=0
        fail_count=0
        wip_count = 0
        blocked_count = 0
        outscoped_count = 0
        deferred_count = 0
        not_run=0
        for i in range(0,list1[0]['resultSize']):
                if 'lastTestResult' in list1[0]['results'][i]:
                    #print('True')
                    tc_status=list1[0]['results'][i]['lastTestResult']['executionStatus']
                    if tc_status == '1':
                        pass_count += 1
                    if tc_status == '2':
                        fail_count += 1
                    if tc_status == '3':
                        wip_count += 1
                    if tc_status == '4':
                        blocked_count += 1
                    if tc_status == '11':
                        outscoped_count += 1
                    if tc_status == '12':
                        deferred_count += 1
                else:
                    not_run+=1

        print('Total TC_COUNT = '+str(list1[0]['resultSize']))
        print('Total Passed = ' + str(pass_count))
        print('Total Failed = ' + str(fail_count))
        print('Total WIP = ' + str(wip_count))
        print('Total BLocked = ' + str(blocked_count))
        print('Total outscoped = ' + str(outscoped_count))
        print('Total deferred = ' + str(deferred_count))
        print('Total not_run = ' + str(not_run))
        return (str(list1[0]['resultSize']),str(pass_count),str(fail_count),str(wip_count),str(blocked_count),str(outscoped_count),str(deferred_count),str(not_run))

    else:
        print("Error in calling Zephyr!!!")
        return 2

if __name__ == '__main__':
    #cwd=Path.cwd()
    #print(cwd)
    #path=str(cwd)+"\\report_input_output.xlsx"
    #print(path)
    path="report_input_output.xlsx"
    wb_obj=openpyxl.load_workbook(path)
    sheet_obj = wb_obj["INPUT"]
    sheet_obj2 = wb_obj["OUTPUT"]
    cell_obj = sheet_obj.cell(row=2, column=1)
    mail_list= str(sheet_obj.cell(row=2, column=3).value).strip()
    print ('Mail_list is - ' +mail_list)
    mail_subject = str(sheet_obj.cell(row=2, column=4).value)
    print('Mail_subject is - ' +mail_subject)
    print ()
    if(mail_list == 'None' or mail_subject == 'None' ):
        print("Mail To list is empty in excel or Subject is empty.")
        sys.exit(0)
    #print(cell_obj.value)
    #print(sheet_obj.max_row)
    m_row = sheet_obj.max_row
    flag=0
    if m_row>0:
        pass_total=0
        not_executed_total=0
        grand_total=0
        #string_html_table=""
        string_html_table = '<p>Hi All,</p><p>TC Count from Zephyr</p><p>&nbsp;</p><table style="border-collapse: collapse; width: 90.4153%; height: 90px;" border="1"><tbody><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px; background-color: #99ccff;"><strong>Testcase Alt ID</strong></td><td style="width: 14.2422%; height: 18px; background-color: #99ccff;"><strong>Pass Count</strong></td><td style="width: 16.272%; height: 18px; background-color: #99ccff;"><strong>Not Executed</strong></td><td style="width: 14.6893%; height: 18px; background-color: #99ccff;"><strong>Total TCs</strong></td></tr>'
        for i in range(2, m_row + 1):
            cell1 = str(sheet_obj.cell(row=i, column=1).value)
            cell2 = str(sheet_obj.cell(row=i, column=2).value)
            #if !(cell1.value == None) & (str(cell1.value).strip() is None)
            if not(cell1.isspace()) and not(cell2.isspace())  and (cell2 != 'None') and (cell1 != 'None') :
                flag=1
                print('\n\nInput release is - '+cell1.strip()+', Input alt_id is - '+cell2.strip())
                output=(get_data_of_alt_id(cell1.strip(),cell2.strip()))
                if output==1:
                    print ("Error in fetching data for release - "+cell1.strip())
                if output==2:
                    print("Error in getting TCs for the given alt_id - "+cell2.strip())
                else:
                    #print(type(output))
                    sheet_obj2.cell(row=i, column=1).value=cell2.strip()
                    sheet_obj2.cell(row=i, column=2).value=cell1.strip()
                    sheet_obj2.cell(row=i, column=3).value = output[0] #total tc
                    sheet_obj2.cell(row=i, column=5).value = str(int(output[1]) + int(output[2])) #total execute
                    sheet_obj2.cell(row=i, column=6).value = output[1] #total passed
                    sheet_obj2.cell(row=i, column=10).value = output[2]  #total failed
                    pass_total=pass_total+int(output[1])
                    not_executed_total=not_executed_total+int(output[7])
                    grand_total = grand_total + int(output[0])
                    string_html_table += '<tr style="height: 18px;"><td style="width: 47.2015%; height: 18px;"><pre>'+cell2.strip()+'</pre></td><td style="width: 14.2422%; height: 18px;"><pre>'+output[1]+'</pre></td><td style="width: 16.272%; height: 18px;"><pre>'+output[7]+'</pre></td><td style="width: 14.6893%; height: 18px;"><pre>'+output[0]+'</pre></td></tr>'
        string_html_table +='<tr style="height: 18px;"><td style="width: 47.2015%; height: 18px; background-color: #99ccff;"><strong>Grand Total</strong></td><td style="width: 14.2422%; height: 18px; background-color: #99ccff;"><pre><strong>'+str(pass_total)+'</strong></pre></td><td style="width: 16.272%; height: 18px; background-color: #99ccff;"><pre><strong>'+str(not_executed_total)+'</strong></pre></td><td style="width: 14.6893%; height: 18px; background-color: #99ccff;"><pre><strong>'+str(grand_total)+'</strong></pre></td></tr></tbody></table><p>&nbsp;</p><p><em>- Auto Generated mail</em></p>'
    if flag==0:
        string_html_table=""
        print("None of the Input release/altid is valid or all are empty. So no report generated, hence no mail")
    else:
        print (string_html_table)
        outlook = win32.Dispatch('outlook.application')  # get a reference to Outlook
        mail = outlook.CreateItem(0)  # create a new mail item
        mail.To = mail_list  #'faizan.khan@amdocs.com;mahjabeen.bano@amdocs.com'
        today = date.today()
        mail.Subject = str(mail_subject) +' Execution Status Report - 02/26'
        mail.HTMLBody = '<p>Hi All,</p><p>TC Count from Zephyr</p><p>&nbsp;</p><table style="border-collapse: collapse; width: 90.4153%; height: 90px;" border="1"><tbody><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px; background-color: #99ccff;"><strong>Testcase Alt ID</strong></td><td style="width: 14.2422%; height: 18px; background-color: #99ccff;"><strong>Pass Count</strong></td><td style="width: 16.272%; height: 18px; background-color: #99ccff;"><strong>Not Executed</strong></td><td style="width: 14.6893%; height: 18px; background-color: #99ccff;"><strong>Total TCs</strong></td></tr><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px;"><pre>2104_QA_DTV_Reg</pre></td><td style="width: 14.2422%; height: 18px;"><pre>20</pre></td><td style="width: 16.272%; height: 18px;"><pre>480</pre></td><td style="width: 14.6893%; height: 18px;"><pre>500</pre></td></tr><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px;"><pre>2104_SIT_Reg</pre></td><td style="width: 14.2422%; height: 18px;"><pre>0</pre></td><td style="width: 16.272%; height: 18px;"><pre>1104</pre></td><td style="width: 14.6893%; height: 18px;"><pre>1104</pre></td></tr><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px;"><pre>2104_UNF_REG</pre></td><td style="width: 14.2422%; height: 18px;"><pre>40</pre></td><td style="width: 16.272%; height: 18px;"><pre>917</pre></td><td style="width: 14.6893%; height: 18px;"><pre>957</pre></td></tr><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px;"><pre>PID400476a_ENB_ON</pre></td><td style="width: 14.2422%; height: 18px;"><pre>5</pre></td><td style="width: 16.272%; height: 18px;"><pre>205</pre></td><td style="width: 14.6893%; height: 18px;"><pre>210</pre></td></tr><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px;"><pre>PID325799_CR207270_ENB</pre></td><td style="width: 14.2422%; height: 18px;"><pre>1</pre></td><td style="width: 16.272%; height: 18px;"><pre>22</pre></td><td style="width: 14.6893%; height: 18px;"><pre>23</pre></td></tr><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px;"><pre>PID400379_ENB_OFF</pre></td><td style="width: 14.2422%; height: 18px;"><pre>5</pre></td><td style="width: 16.272%; height: 18px;"><pre>77</pre></td><td style="width: 14.6893%; height: 18px;"><pre>82</pre></td></tr><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px;"><pre>PID400379_ENB_ON</pre></td><td style="width: 14.2422%; height: 18px;"><pre>7</pre></td><td style="width: 16.272%; height: 18px;"><pre>173</pre></td><td style="width: 14.6893%; height: 18px;"><pre>180</pre></td></tr><tr style="height: 18px;"><td style="width: 47.2015%; height: 18px; background-color: #99ccff;"><strong>Grand Total</strong></td><td style="width: 14.2422%; height: 18px; background-color: #99ccff;"><pre><strong>78</strong></pre></td><td style="width: 16.272%; height: 18px; background-color: #99ccff;"><pre><strong>2978</strong></pre></td><td style="width: 14.6893%; height: 18px; background-color: #99ccff;"><pre><strong>3056</strong></pre></td></tr></tbody></table><p>&nbsp;</p><p><em>- Auto Generated mail</em></p>'
        mail.Send()
        
    wb_obj.save(path)
    wb_obj.close()
    #print(pass_total, not_executed_total,grand_total )
